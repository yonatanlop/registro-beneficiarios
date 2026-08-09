# -*- coding: utf-8 -*-
"""
Genera data/transcripcion_junio_2025.xlsx a partir de la lectura manual
(asistida por IA) de las 11 paginas escaneadas de
"Planillas Jornada de apoyo junio 2025.pdf".

Este archivo es un BORRADOR PARA REVISION HUMANA: la letra manuscrita del
PDF puede generar errores de lectura, sobre todo en numeros de documento y
telefono. Antes de importarlo en la aplicacion, alguien del equipo debe
revisar cada fila (especialmente las marcadas en la columna
"Revisar (IA)") comparando contra el PDF original.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Datos comunes de la jornada (iguales en las 11 paginas del PDF)
JORNADA = {
    "fecha": "",  # el campo "3. FECHA DILIGENCIAMIENTO" quedo en blanco en el PDF original
    "departamento": "BOYACA",
    "municipio": "TUNJA",
    "lugar": "CRA 10 N 19 - 21 AUDIT. EDUARDO CALDERON",
    "actividad": "ENTREGA DE MERCADOS BASICOS",
    "linea": "AYUDANOS A AYUDAR",
    "programa": "JORNADA DE APOYO",
    "proyecto": "FAMILIAS VULNERABLES",
    "beneficios": "ENTREGA DE MERCADOS",
    "familias": True,
    "adulto_mayor": False,
    "ninez": False,
    "etnias": False,
}

# (pagina, fila, nombres, documento, sexo, edad, disc[aud,vis,cog,men,fis],
#  direccion, zona, telefono, representado, observaciones)
FILAS = [
(1,1,"Nelcy Cardona","41108004","F",47,"","Calle 5 Sur # 9-34","La Perla","3214974052",False,"Numero de calle poco legible"),
(1,2,"Dagny Urrego","68302897","F",50,"","Cra 12 # 7-86","Libertador","3166622673",False,""),
(1,3,"Maria Consuelo Romero","24017660","F",55,"","Calle 4B Sur # 7-63","San Francisco","3102268314",False,"Apellido poco legible"),
(1,4,"Angie Maritza Florez Bautista","1049641146","F",30,"","Vereda Runta Arriba, Tunja","","3108195602",False,""),
(1,5,"Maria Jimenez Bohorquez","24156924","F",54,"","Mz 3 Bloque 4 Apto 304","Antonia Santos","3103329200",False,""),
(1,6,"Leticia Patarroyo Patino","51782895","F",59,"","Cra 10 # 14-88","San Laureano","3228981348",False,""),
(1,7,"Doris Stella Casallas","52491130","F",47,"","Calle 15A # 12B-04","Nueva Santa Barbara","3114645800",False,""),
(1,8,"Yenny Carolina Ibanez Lopez","1049629514","F",33,"","Calle 25 # 16-04","Carmen","3115044809",False,""),
(1,9,"Elva Carolina Saenz Hernandez","23315701","F",53,"VIS","Mz 6 Torre 2 Apto 503","Antonia Santos","3203892751",False,""),
(1,10,"Elsa Ruth Gomez Cuadros","52008392","F",56,"","Cra 21 # 6A-93","San Ignacio","3203264620",False,""),

(2,1,"Ana D. Rojas Ortega","23268924","F",80,"","Cra 12 # 2A-25","Surinama","3192600189",False,""),
(2,2,"Jose Del Carmen Sanchez","6762951","M",67,"","Av Norte 37-22","Glorieta","3012653771",False,""),
(2,3,"Gloria I. Bustos","41773165","F",67,"","Clle 7A # 12-39","Libertador","3102110942",False,""),
(2,4,"Poliadro Cardenas","4149779","M",72,"","","San Lazaro","3208189343",False,"Direccion no diligenciada en el PDF"),
(2,5,"Aurelia Cifuentes T.","41732576","F",66,"","Cra 12 # 25-37","San Lucia","3115444207",False,""),
(2,6,"Martha E. Moreno","51928796","F",59,"","Clle 11 # 11-70","Aquimin","3118889405",False,""),
(2,7,"Luz Marina Cuadros","2398330","F",69,"","Clle 11C # 10A-16","Aquimin","3102536895",False,""),
(2,8,"Maricela Benavides","24070997","F",40,"","Av Progreso 9","Monteverde","3138236367",False,""),
(2,9,"Wilmer Berdugo","1057588487","M",33,"","Trs 2E 65-66","Muiscas","3232486426",False,""),
(2,10,"Luz Janeth Jerez","30206205","F",50,"","Dcra 62 # 8-45","Asis","3115732038",False,"Telefono: primer digito ambiguo (3 u 8)"),

(3,1,"Jakelin Villalobos","1002565451","F",21,"","Cra 5 # 8-02","Jordan","3203934680",False,"Telefono con un digito ilegible, verificar"),
(3,2,"Yudy L. Porras S.","46385207","F",41,"","Cra 11B # 13-82","Bosque","3027193419",False,""),
(3,3,"Sildana Quinonez","52497537","F",46,"","Cra 14A # 55B-16","M. Carlo","3174558172",False,""),
(3,4,"Laura J. Rojas G.","1049658900","F",25,"","Clle 28 # 20A-35","Altamira","3208204433",False,""),
(3,5,"Rosa E. Espinosa","85406102","F",63,"","Manzana I # 7","Florida","3214175643",False,""),
(3,6,"Maria M. Garcia S.","40024334","F",58,"","Clle 28 # 20A-13","A. Mira","3125494184",False,""),
(3,7,"Ninfa Jimenez","34564450","F",58,"","Vda Virgua","T. Virgua","3114819873",False,""),
(3,8,"Juan M. Bonilla","71180131","M",45,"COG","Cra 12 # 7B-15","Libertador","3103267975",True,""),
(3,9,"Jose Granados","17023454","M",85,"","Clle 8N B-34","Suarez","3102202668",True,""),
(3,10,"Ines Nino","40015016","F",68,"","Tras Alto Man","El Alto","3135118052",False,""),

(4,1,"Martha Lucia Echavarria","40034591","F",69,"AUD","Calle 7A # 12-39","Libertador","3134905527",False,""),
(4,2,"Yuli Amanda Florez","1049605329","F",37,"","Vereda Runta","Runta","3108706596",False,""),
(4,3,"Celmira Socadagui","40014244","F",65,"","Cra 12 # 6-30","Libertador","3132830939",False,"Apellido poco legible"),
(4,4,"Luzmarina Cunarado","41744871","F",66,"","Calle 12 # 5-21","Jordan","3208741582",False,"Apellido poco legible"),
(4,5,"Demy Piratova S.","7188305","M",40,"","Torre 5E","Antonia Santos","3184922268",False,""),
(4,6,"Blanca Irene Cabo L.","40045812","F",51,"","Mz 13 Casa 15","Pinos de Oriente","3114955822",False,""),
(4,7,"Dory Andrea Celyn M.","40040080","F",49,"","Calle 166 # 51-22","Colinas del Norte","3102038495",False,"Nombre/direccion poco legibles"),
(4,8,"Leonelda Suarez","40031965","F",54,"","Cra 1C # 18-62","Prados de Alcala","3114621534",False,""),
(4,9,"Daniel Rubio","1050602853","M",20,"","Calle 13 # 17-55","Paraiso","3242113745",False,""),
(4,10,"Delia Rosero","40037093","F",51,"","Calle 18 # 6-88","San Ignacio","3125862857",False,""),

(5,1,"Lida Del Carmen Cano H.","40022804","F",59,"","Cra 12 # 54A-23","Jose Antonio Galan","3125799261",False,""),
(5,2,"Benilda Suspes P.","20340683","F",81,"AUD","Vereda Combita","San Rafael","3222482872",False,""),
(5,3,"Rosalbina Moreno","40011846","F",65,"","Clle 12 # 12-73 Int 2","Americas","3214503150",False,""),
(5,4,"Luis Alejandro Cortes S.","17056069","M",83,"","Cra 2A # 42A-69","Las Quintas","3204590695",False,""),
(5,5,"Alba Luz Rojas","46680739","F",53,"","Vereda Forantiva, Oicata","","3124370875",False,""),
(5,6,"Ascencion Munoz","24196981","F",65,"","Cra 5 # 18-90","San Ignacio","3123204852",False,""),
(5,7,"Ana Teresa Martinez","29897401","F",77,"","Diag 67 # 0-10 Int 1","Los Muiscas","3156778790",False,""),
(5,8,"Edelmira Lopez","40013339","F",69,"","Calle 74 # 5-30","Palos Verdes","3213460354",False,""),
(5,9,"Luz Yaneth Rodriguez","23490091","F",75,"","Calle 4 # 4-80","Cooservicios","3124104791",False,""),
(5,10,"Aracelly Suarez P.","39403708","F",62,"","Calle 31 # 13-25","Gaitan","3213177469",False,""),

(6,1,"Doris S. Yaruquen","1002392949","F",36,"","Vd Runta Bajo","Runta","3167683905",False,""),
(6,2,"Elensy Tirado","32272262","F",46,"","Clle 7 # 12-69","Libertador","3204990591",False,""),
(6,3,"Adriana Salazar","40037454","F",51,"","Clle 24 # 9-100","Nieves","3228083044",False,""),
(6,4,"Luz Marina Quitian","23157411","F",50,"","Bloque 11","B. Antonia Santos","3168034339",False,""),
(6,5,"Olga Alvarez","46668200","F",52,"","Diag 60 # 7-02","V. Luz","3102782922",False,""),
(6,6,"Rudy Florez","1049628231","F",33,"","Vda Runta","Runta","3204295139",False,""),
(6,7,"Paola Baron L.","1057606134","F",27,"","Clle 8N 10-30","Suarez A.","3212508275",False,""),
(6,8,"Maritza Guerrero L.","24178941","F",53,"","Cra 15H 24-45","San Lucia","3124125528",False,""),
(6,9,"Flavio Ausaque","4046379","M",59,"","Vda Puente Hamaca","Soraca","3102390018",False,"Telefono: primer digito ambiguo (3 u 8)"),
(6,10,"Gloria Doncel","23399844","F",48,"","Clle 2 # 5A-27","A. Santos","3142706610",False,""),

(7,1,"Clara Ines Quijano Molina","28680293","F",69,"AUD;FIS","Barrio Antonia Santos Torre 84 Apto 103","Antonia Santos","3132710314",False,""),
(7,2,"Maria Ilda Jimenez de Munoz","23636509","F",71,"AUD","Cra 8a # 2-14","Antonia Santos","3142011272",False,""),
(7,3,"Angela Maria Gama Rojas","1055186226","F",32,"","Vereda Forantiva, Oicata","","3186494442",False,""),
(7,4,"Yirley Monjarres","45540192","F",42,"","Calle 11A # 5-15","Jordan","3116022048",False,""),
(7,5,"Adriana Lucia Cuervo Amaya","1002709687","F",23,"","Siachoque","Primavera","3236034999",False,""),
(7,6,"Custodia Sanchez Rincon","40018647","F",64,"","Torre 5E Apto 302","Antonia Santos","3142980678",False,""),
(7,7,"Paola Florez","1030566809","F",35,"","Cra 9a # 6-10","Obrero","3102498721",False,""),
(7,8,"Luz Mila Medina Quinchara","41618018","F",72,"","Cra 2a # 2-64 Casa 34","Combita","3108142682",False,""),
(7,9,"Gloria Casteblanco Rativa","40036364","F",53,"","Calle 15A # 15-76","Ricaurte","3212699909",False,""),
(7,10,"Martha Patricia Cortes Blanco","40023457","F",60,"","Cra 2B Este 11-06 Sur","Ciudad Jardin","3115947094",False,""),

(8,1,"Ana Belen Yanguen Rivera","23266302","F",77,"","Cra 5 Este # 27-44","El Dorado","3168964605",False,""),
(8,2,"Maria Teresa Santamaria","28476751","F",76,"","Clle 7 # 19-46","Mirador Escandinavo","3186236965",False,""),
(8,3,"Pedro Jaime Zipa Moapllon","6757234","M",70,"","Clle 11 # 8-25","San Laureano","3206404695",False,"Apellido poco legible"),
(8,4,"David Largo Cadena","1049618117","M",35,"","Clle 7C # 19-39","Mirador Escandinavo","3136887121",False,""),
(8,5,"Maria Berenice Vargas Monroy","41732526","F",69,"","Cra 21 # 31-85","Altamira","3226094216",False,""),
(8,6,"Marceliano Camargo","6746054","M",81,"","Cra 18 # 76-41","Mirador Escandinavo","3133946758",False,""),
(8,7,"Blanca Cecilia Cruz H.","40028978","F",57,"","Mz 9 Casa 22","Mirador Escandinavo","3142911157",False,""),
(8,8,"Nidia Stella Rojas Lopez","40034853","F",53,"","Cra 4A # 3-23","Ciudadela Sol de Oriente","3208803132",False,""),
(8,9,"Nubia Consuelo Reyes Suarez","40045586","F",49,"","Cra 17A # 8-09","Colinas de San Fernando","3118806355",False,""),
(8,10,"Aura Cecilia Silva Sanchospe","40012619","F",68,"","Clle 50C # 9F-29","Estancia del Roble","3102281800",False,"Apellido poco legible"),

(9,1,"Nancy Cecilia Barrera Ledesma","52073832","F",54,"","Clle 15A # 1-20","Patriotas","3118057016",False,""),
(9,2,"Gloria Esperanza Jara Novoa","40026258","F",57,"","Clle 4 # 12-47","Libertador","3195188735",False,""),
(9,3,"Maria Carmenza Cruz Bohorquez","40032930","F",53,"","Dg 6B Sur # 5-08","San Francisco","3123575650",False,""),
(9,4,"Nalciris Rodriguez Munoz","28313516","F",55,"","Cr 2 # 15-32","Patriotas","3107553635",False,"Nombre poco legible"),
(9,5,"Flor Myriam Barrios Celis","28892572","F",69,"","Cr 3A # 22","Los Pinitos","3017574047",False,""),
(9,6,"Jose Francisco Gil Molina","6771995","M",60,"","Clle 30 # 14-29","20 de Julio","3206832327",False,""),
(9,7,"Ana Milena Morales Camacho","23756003","F",43,"","Clle 8 # 10-30","Suarez","3123253118",False,""),
(9,8,"Ingrid Lorena Reyes Daza","1050603009","F",20,"","Clle 2A # 15-03","Bolivar","3135159873",False,""),
(9,9,"Argie Katherine Munoz Aponte","1049656336","F",26,"","Cra 13 # 8-15 Sur","San Carlos","3106190633",False,""),
(9,10,"Maria Del P. Barajas","40049010","F",44,"FIS","Diag 68A # 02-45","Colinas de Ture","3125443831",False,"Zona/barrio poco legible"),

(10,1,"Nahir C. Martinez P.","1053282379","F",30,"","Calle 3 # 16-68","Triunfo","3182110993",False,""),
(10,2,"Ana Joaquina Munoz","51956255","F",55,"","Cra 8 # 12-14","Antonia Santos","3114993044",False,""),
(10,3,"Miryam Leonor Rojas C.","40011117","F",68,"","Cra 33 # 32-33","Ciudadela Confaboy","3118955498",False,""),
(10,4,"Valeria M. Larrota","1051066181","F",20,"","Calle 3a Sur # 15-30","Trinidad","3209627477",False,""),
(10,5,"Estiven D. Suarez G.","1050090134","M",20,"","Tras del Alto","Mancan Arriba","3246019679",False,""),

(11,1,"Julio Hernando Castro","74341892","M",44,"","Calle 16 # 17-92","Ricaurte","3112760712",False,""),
(11,2,"Jessica Vanesa Gaitan Carrero","1020823471","F",28,"","Cra 12 # 7B-08/14","Libertador","3208862063",False,""),
]

HEADERS = [
    "Pagina PDF", "Fila en planilla",
    "Fecha de diligenciamiento", "Departamento", "Municipio", "Lugar y dirección",
    "Actividad desarrollada", "Línea estratégica", "Programa", "Proyecto",
    "Nombre de los beneficios entregados", "Familias", "Adulto mayor", "Niñez", "Etnias",
    "Nombres y apellidos completos", "Documento de identidad", "Sexo", "Edad",
    "Auditiva", "Visual", "Cognitiva", "Mental", "Física",
    "Dirección de residencia", "Zona o barrio", "Teléfono", "Representado",
    "Revisar (IA)",
]

DISC_CODE = {"AUD": 0, "VIS": 1, "COG": 2, "MEN": 3, "FIS": 4}


def disc_flags(code_str):
    flags = [False, False, False, False, False]
    for part in code_str.split(";"):
        part = part.strip()
        if part in DISC_CODE:
            flags[DISC_CODE[part]] = True
    return flags


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="FFDCE6F1")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in FILAS:
        pagina, fila, nombres, doc, sexo, edad, disc, direccion, zona, tel, repr_, obs = r
        aud, vis, cog, men, fis = disc_flags(disc)
        ws.append([
            pagina, fila,
            JORNADA["fecha"], JORNADA["departamento"], JORNADA["municipio"], JORNADA["lugar"],
            JORNADA["actividad"], JORNADA["linea"], JORNADA["programa"], JORNADA["proyecto"],
            JORNADA["beneficios"],
            "X" if JORNADA["familias"] else "", "X" if JORNADA["adulto_mayor"] else "",
            "X" if JORNADA["ninez"] else "", "X" if JORNADA["etnias"] else "",
            nombres, doc, sexo, edad,
            "X" if aud else "", "X" if vis else "", "X" if cog else "", "X" if men else "", "X" if fis else "",
            direccion, zona, tel,
            "X" if repr_ else "",
            obs,
        ])

    widths = [10, 10, 14, 14, 12, 30, 22, 18, 16, 18, 22, 10, 12, 8, 8,
              30, 16, 8, 8, 9, 9, 9, 9, 9, 30, 20, 14, 12, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(FILAS)+1}"

    out_path = "data/transcripcion_junio_2025.xlsx"
    wb.save(out_path)
    print(f"Guardado {out_path} con {len(FILAS)} filas")


if __name__ == "__main__":
    build()
